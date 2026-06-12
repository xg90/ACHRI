from __future__ import annotations

import re
import os
from pathlib import Path

import numpy as np
from PIL import Image
import openpyxl

# -------- Config --------
SCRIPT_DIR = Path(__file__).resolve().parent
# Use current working directory, then join target folders by name.
CWD = Path(os.getcwd())
BASE = CWD / "AC&HCI" / "project"

PPG1_DIR = BASE / "PPG1"
PPG2_DIR = BASE / "PPG2"
FACES_DIR = BASE / "aligned_faces_64"
LABEL_PATH = BASE / "label.xlsx"

OUTPUT_DIR = CWD

N_SUBJECTS = 16
N_SECONDS = 120
PPG_HZ = 250
FRAMES_PER_SEC = 16
STRICT = True  # True: stop on missing/insufficient data; False: skip

# "by_recording": batch1->PPG1 (recording1), batch2->PPG2 (recording2)
# "both": stack PPG1 + PPG2 as 2 channels
PPG_MODE = "by_recording"
# ------------------------


def read_ppg_xlsx(path: Path) -> np.ndarray:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # skip header
        v = row[0] if row else None
        if v is None:
            continue
        values.append(float(v))
    return np.asarray(values, dtype=np.float32)


def ensure_length(arr: np.ndarray, length: int, name: str) -> np.ndarray:
    if arr.size < length:
        raise ValueError(f"{name} has {arr.size} samples, expected at least {length}.")
    if arr.size > length:
        arr = arr[:length]
    return arr


def sort_frames(paths):
    def key_fn(p: Path):
        m = re.search(r"(\\d+)", p.stem)
        return int(m.group(1)) if m else p.stem
    return sorted(paths, key=key_fn)


def select_frames(paths, target_count: int):
    if len(paths) < target_count:
        return None
    if len(paths) == target_count:
        return paths
    idx = np.linspace(0, len(paths) - 1, target_count)
    idx = np.round(idx).astype(int)
    idx[0] = 0
    idx[-1] = len(paths) - 1
    return [paths[i] for i in idx]


def load_faces(sim_dir: Path) -> np.ndarray | None:
    frame_paths = sort_frames(list(sim_dir.glob("*.jpg")))
    frame_paths = select_frames(frame_paths, N_SECONDS * FRAMES_PER_SEC)
    if frame_paths is None:
        return None

    frames = []
    for p in frame_paths:
        img = Image.open(p).convert("RGB")
        if img.size != (64, 64):
            img = img.resize((64, 64), Image.BILINEAR)
        arr = np.asarray(img, dtype=np.float32) / 255.0
        arr = np.transpose(arr, (2, 0, 1))  # (3, 64, 64)
        frames.append(arr)

    frames = np.stack(frames, axis=0)  # (T, 3, 64, 64)
    frames = frames.reshape(N_SECONDS, FRAMES_PER_SEC, 3, 64, 64)
    return frames


def build_ppg(subject_id: int, recording: str) -> np.ndarray:
    ppg1 = read_ppg_xlsx(PPG1_DIR / f"{subject_id}.xlsx")
    ppg2 = read_ppg_xlsx(PPG2_DIR / f"{subject_id}.xlsx")

    ppg1 = ensure_length(ppg1, N_SECONDS * PPG_HZ, f"PPG1/{subject_id}")
    ppg2 = ensure_length(ppg2, N_SECONDS * PPG_HZ, f"PPG2/{subject_id}")

    ppg1 = ppg1.reshape(N_SECONDS, PPG_HZ)
    ppg2 = ppg2.reshape(N_SECONDS, PPG_HZ)

    if PPG_MODE == "by_recording":
        if recording == "recording1":
            return ppg1[:, None, :]
        return ppg2[:, None, :]
    return np.stack([ppg1, ppg2], axis=1)  # (120, 2, 250)


def read_labels(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    batch1 = {}
    batch2 = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        c1, l1, _, c2, l2 = row
        if c1 is not None:
            batch1[int(c1)] = int(l1)
        if c2 is not None:
            batch2[int(c2)] = int(l2)

    return batch1, batch2


def build_batch(recording: str, label_map: dict[int, int], out_name: str) -> None:
    ppg_list = []
    faces_list = []
    labels_list = []
    subject_ids = []

    for subject_id in range(1, N_SUBJECTS + 1):
        sim_dir = FACES_DIR / recording / f"video_{subject_id}" / "simulate"
        if not sim_dir.exists():
            msg = f"Missing face dir: {sim_dir}"
            if STRICT:
                raise FileNotFoundError(msg)
            print(f"Skip {recording} sub{subject_id:02d}: {msg}")
            continue

        if subject_id not in label_map:
            msg = f"Missing label for clip {subject_id} in {recording}"
            if STRICT:
                raise ValueError(msg)
            print(f"Skip {recording} sub{subject_id:02d}: {msg}")
            continue

        faces = load_faces(sim_dir)
        if faces is None:
            msg = "not enough frames (<1920)"
            if STRICT:
                raise ValueError(f"{recording} sub{subject_id:02d}: {msg}")
            print(f"Skip {recording} sub{subject_id:02d}: {msg}")
            continue

        ppg = build_ppg(subject_id, recording)
        label = label_map[subject_id]

        ppg_list.append(ppg)
        faces_list.append(faces)
        labels_list.append(np.full((N_SECONDS,), label, dtype=np.int64))
        subject_ids.append(np.full((N_SECONDS,), subject_id, dtype=np.int64))

    if not ppg_list:
        raise RuntimeError(f"No data found for {recording}")

    ppg_all = np.concatenate(ppg_list, axis=0)       # (N*120, C, 250)
    faces_all = np.concatenate(faces_list, axis=0)   # (N*120, 16, 3, 64, 64)
    labels_all = np.concatenate(labels_list, axis=0) # (N*120,)
    subject_all = np.concatenate(subject_ids, axis=0)

    out_path = OUTPUT_DIR / out_name
    np.savez_compressed(
        out_path,
        ppg=ppg_all,
        faces=faces_all,
        labels=labels_all,
        clip_id=subject_all,
    )
    print(
        f"Saved {out_path} | ppg {ppg_all.shape} | faces {faces_all.shape} | labels {labels_all.shape}"
    )


def main() -> None:
    batch1_labels, batch2_labels = read_labels(LABEL_PATH)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    build_batch("recording1", batch1_labels, "BATCH1.npz")
    build_batch("recording2", batch2_labels, "BATCH2.npz")


if __name__ == "__main__":
    main()
